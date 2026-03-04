import pyautogui as p
import sys
from time import sleep as t
import cv2
import re
import numpy as np
from PIL import Image
import pytesseract
import os
import random
##import EOB_data

# loop virtual punch fly wine answer wheel equal artist stand mosquito cook
p.PAUSE = 0.005


def ocrImageDetectionUI(x1, y1, x2 , y2, variableText, variableText1):
    print("please wait extracting text")
    global text
    global OCR_value
    global text_array
    global all_num_value
    all_num_value = False
    text_array = []
    global all_num
    OCR_value = False
    all_num = ""
    global xAxis, yAxis
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

    screenshot = p.screenshot()
    screenshot.save(main_path + r"\screenshot.png")
    bigImage = (main_path + r"\screenshot.png")
    
    img = cv2.imread(bigImage)    #(r'C:\Users\aman\Desktop\ocrCode\screenshot.png')
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

    global image
    image = Image.open(bigImage) 

    image_arr = np.array(image) 
    image_arr = image_arr[y1:y2, x1:x2]
    image = Image.fromarray(image_arr)
    image.save(main_path + r"\small_screenshot.png")
    smallImg = image.save(main_path + r"\small_screenshot.png")
    OCR_value = False

    smallImg = cv2.imread(r"C:\Users\aman\Desktop\small_screenshot.png")
    img = cv2.cvtColor(smallImg, cv2.COLOR_BGR2GRAY)
    sharpen_kernel = np.array([[-1,-1,-1], [-1,9,-1], [-1,-1,-1]])
    sharpen = cv2.filter2D(img, -1, sharpen_kernel)
    thresh = cv2.threshold(sharpen, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
    
#def OCRFunction(image):
    boxes = pytesseract.image_to_data(thresh, lang='eng', config='--psm 6')
    ##print(boxes)
    a = 0
    for x, b in enumerate(boxes.splitlines()):
        if x!=0:
            b = b.split()
            if len(b)==12:
                a = a + 1
                text = str(b[11])
                text_array.append(text)
                y = re.findall(r"\d", text)
                all_num = "".join(y)
                # OCR_value = False
                print(text)
                if(variableText in text or variableText1 in text):
                    xAxis = int(b[6]) + x1
                    yAxis = int(b[7]) + y1
                    OCR_value = True
                    print(xAxis, ',' , yAxis)
                    return False
##                print(text)
                # OCR_value = False
                if(all_num != ""):
                    xAxis = int(b[6]) + x1
                    yAxis = int(b[7]) + y1
                    all_num_value = True
                    # print(all_num)
                    print(xAxis, ',' , yAxis)
                    return False
            a = a + 1
    return OCR_value
    return xAxis
    return yAxis
    return text
    return all_num
    return all_num_value

        
def ocrImageDetection(x1, y1, x2 , y2, variableText, variableText1):
    print("please wait extracting text")
    global text
    global text_array
    text_array = []
    global OCR_value
    OCR_value = False
    global xAxis, yAxis
    pytesseract.pytesseract.tesseract_cmd = r'C:\Users\aman\Desktop\ocrCode\Tesseract-OCR\tesseract.exe'

    screenshot = p.screenshot()
    screenshot.save(main_path + r"\screenshot.png")
    bigImage = (main_path + r"\screenshot.png")
    
    img = cv2.imread(bigImage)    #(r'C:\Users\aman\Desktop\ocrCode\screenshot.png')
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

    global image
    image = Image.open(bigImage) 

    image_arr = np.array(image) 
    image_arr = image_arr[y1:y2, x1:x2]
    image = Image.fromarray(image_arr)
    OCR_value = False
#def OCRFunction(image):
    boxes = pytesseract.image_to_data(image, lang='eng', config='--psm 3')      #put psm 7 if this fails.
    ##print(boxes)
    a = 0
    for x, b in enumerate(boxes.splitlines()):
        if x!=0:
            b = b.split()
            if len(b)==12:
                a = a + 1
                text = str(b[11])
                print(text)
                text_array.append([text, int(b[6]) + x1, int(b[7]) + y1])
                # OCR_value = False
                if(variableText in text or variableText1 in text):
                    xAxis = int(b[6]) + x1
                    yAxis = int(b[7]) + y1
                    OCR_value = True
                    print(xAxis, ',' , yAxis)
                    return False
            a = a + 1
    return OCR_value
    return xAxis
    return yAxis
    return text
    return text_array

def continous_ocrImageDetection(x1, y1, x2 , y2, variableText, variableText1):
    print("please wait extracting text")
    global text
    global complete_text_array
    complete_text_array = []
    global OCR_value
    OCR_value = False
    global xAxis, yAxis
    pytesseract.pytesseract.tesseract_cmd = r'C:\Users\aman\Desktop\ocrCode\Tesseract-OCR\tesseract.exe'

    screenshot = p.screenshot()
    screenshot.save(main_path + r"\screenshot.png")
    bigImage = (main_path + r"\screenshot.png")
    
    img = cv2.imread(bigImage)    #(r'C:\Users\aman\Desktop\ocrCode\screenshot.png')
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

    global image
    image = Image.open(bigImage) 

    image_arr = np.array(image) 
    image_arr = image_arr[y1:y2, x1:x2]
    image = Image.fromarray(image_arr)
    OCR_value = False
#def OCRFunction(image):
    boxes = pytesseract.image_to_data(image, lang='eng', config='--psm 3')      #put psm 7 if this fails.
    ##print(boxes)
    a = 0
    for x, b in enumerate(boxes.splitlines()):
        if x!=0:
            b = b.split()
            if len(b)==12:
                a = a + 1
                text = str(b[11])
                print(text)
                complete_text_array.append([text, int(b[6]) + x1, int(b[7]) + y1])
                # OCR_value = False
                if(variableText in text or variableText1 in text):
                    xAxis = int(b[6]) + x1
                    yAxis = int(b[7]) + y1
                    OCR_value = True
                    print(xAxis, ',' , yAxis)
            a = a + 1
    return OCR_value
    return xAxis
    return yAxis
    return text
    return complete_text_array


def ocrImageDetection_delimeter(xAxis, yAxis, delimiterx, delimitery, variableText, bigImage):
    #mainImg, 
    # delimiterx, delimitery = 25, 25
    x1, y1, x2, y2 = xAxis - int(delimiterx), yAxis - int(delimitery), xAxis + int(delimiterx), yAxis + int(delimitery)
    # #print(x1, y1, x2, y2)
    if(x1 <= 0):
       x1 ,x2 = 0, x2 + -(x1)
    if(y1 <= 0):
       y1 ,y2 = 0, y2 + -(y1)
    if(x2 <= 0):
       x2, x1 = 0, x1 + -(x2)
    if(y2 <= 0):
       y2, y1  = 0, y1 + -(y2)
    if(x1 > 1919):
       x2, x1 = -(x1 - 1919 - x2), 1919
    if(y1 > 1079 ):
       y2, y1 = -(y1 - 1079 - y2), 1079
    if(x2 > 1919):
       x1, x2 = -(x2 - 1919 - x1), 1919
    if(y2 > 1079 ):
       y1, y2 = -(y2 - 1079 - y1), 1079
    # #print(x1, y1, x2, y2)
    # img = cv2.imread(mainImg)
    
    print("please wait extracting text")
    global text
    global OCR_value
    OCR_value = False
    pytesseract.pytesseract.tesseract_cmd = r'C:\Users\aman\Desktop\ocrCode\Tesseract-OCR\tesseract.exe'

    img = cv2.imread(bigImage)    #(r'C:\Users\aman\Desktop\ocrCode\screenshot.png')
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

    global image
    image = Image.open(bigImage) 

    image_arr = np.array(image) 
    image_arr = image_arr[y1:y2, x1:x2]
    image = Image.fromarray(image_arr)

#def OCRFunction(image):
    boxes = pytesseract.image_to_data(image)
    ##print(boxes)
    a = 0
    for x, b in enumerate(boxes.splitlines()):
        if x!=0:
            b = b.split()
            if len(b)==12:
                a = a + 1
                text = str(b[11])
                print(text)
                if(variableText in text):
                    xAxis = int(b[6]) + x1
                    yAxis = int(b[7]) + y1
                    OCR_value = True
            a = a + 1
    return OCR_value
    return text


main_path = r"C:\Users\aman\Desktop"




def image_cropper(screenshot_path, xAxis, yAxis, delimiterx, delimitery):
    #mainImg,
    # delimiterx, delimitery = 25, 25
    x1, y1, x2, y2 = xAxis - int(delimiterx), yAxis - int(delimitery), xAxis + int(delimiterx), yAxis + int(delimitery)
    # #print(x1, y1, x2, y2)
    if(x1 <= 0):
       x1 ,x2 = 0, x2 + -(x1)
    if(y1 <= 0):
       y1 ,y2 = 0, y2 + -(y1)
    if(x2 <= 0):
       x2, x1 = 0, x1 + -(x2)
    if(y2 <= 0):
       y2, y1  = 0, y1 + -(y2)
    if(x1 > 1919):
       x2, x1 = -(x1 - 1919 - x2), 1919
    if(y1 > 1079 ):
       y2, y1 = -(y1 - 1079 - y2), 1079
    if(x2 > 1919):
       x1, x2 = -(x2 - 1919 - x1), 1919
    if(y2 > 1079 ):
       y1, y2 = -(y2 - 1079 - y1), 1079
    # #print(x1, y1, x2, y2)
    # img = cv2.imread(mainImg)
    img = cv2.imread(screenshot_path)

    image_arr = np.array(img)      
    # Crop image
    image_arr = image_arr[y1:y2, x1:x2]    #250:350, 735:1174
    # Convert array to image
    image = Image.fromarray(image_arr)
    image.save(screenshot_path)
    # Resizing the image

def imageIdentifier(codeName, screenshot_path, image_path, comparisonInt, xAxis, yAxis, delimiterx, delimitery):
    global comparisonNumber
    global error
    global error_value
    error_value = False
    comparisonNumber = comparisonInt
    screenshot = p.screenshot()
    screenshot.save(screenshot_path)
    image_cropper(screenshot_path, xAxis, yAxis, delimiterx, delimitery)
    # load the input images
    img2 = cv2.imread(image_path)
    img1 = cv2.imread(screenshot_path)
    # convert the images to grayscale
    img1 = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY)
    img2 = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY)
    # define the function to compute MSE between two images
    def mse(img1, img2):
        h, w = img1.shape
        diff = cv2.subtract(img1, img2)
        err = np.sum(diff**2)
        mse = err/(float(h*w))
        return mse, diff
    error, diff = mse(img1, img2)
    if(error <= comparisonInt):
        error_value = True
    else:
        error_value = False
        print("Image matching Error between the two images:",error, "\nHere's the code Name that's showing more errors", codeName)
    return error_value


## Code starts here onwards ##
def cancel_everything_start_from_begining():
    print('cancel_everything_start_from_begining()')

def upper(numb, xAxis, yAxis):
    a = 0
    p.click(x=xAxis, y=yAxis)
    while a < numb:
        p.press('up')
        a = a + 1

def lower(numb, xAxis, yAxis):
    a = 0
    p.click(x=xAxis, y=yAxis)
    while a < numb:
        p.press('down')
        a = a + 1
       
       
def tool_kit():
### Work on it make the popup of twitter disable or work on it do something.
    p.click(button='middle',x=396, y=21)
    t(1)
    p.click(button='left',x=143, y=16)
    ocrImageDetection(x1,y1,x2,y2, "FOLLOW", 'NOODLEHEAD')
    if(OCR_value == True):
        p.click(button='left',x=xAxis, y=yAxis)
        #p.moveTo(x=xAxis, y=yAxis)

### To use icon identifier functions. ###
    patient_name = ["",""]
    main_image_path = r"C:\Users\user\Desktop\python\small_images\main_img.png"
    screenshot_path = r"C:\Users\user\Desktop\python\small_images\small_screenshot.png"
    template_img_path = r"C:\Users\user\Desktop\python\small_images\RPM_approve_icon.png"
    ocrImageDetection(505,226,758,691,patient_name[0],patient_name[1])
    #this will get the patient name and see if my patient is one of them or not. if yes, hover on it
    if(OCR_value == True):
        x1 = 517
        x2 = 1780
        y1 = yAxis - 25
        y2 = yAxis + 25
        image_cropper(main_image_path, x1, y1, x2, y2)
        template_matching(main_image_path, template_img_path, x1, y1, x2, y2)
        p.click(button='left',x=xAxis, y=yAxis)

def clear_python_file():
    with open(r"C:\Users\Amanjyot\Desktop\python\eob_reader\data_store.py", "w") as file_object:
        file_object.write("")
        

def start_writing_data(variableName, text_content_add):
    with open(r"C:\Users\Amanjyot\Desktop\python\eob_reader\data_store.py", "a") as file_object:
        if not text_content_add.strip():
            print(f"{variableName} IS empty")
            text_content_add = ""
            file_object.write(f"{variableName} = '{text_content_add}'\n\n")
        else:
            text_content_add = text_content_add.strip()
            text_content_add = text_content_add.replace("\n", "")
            file_object.write(f"{variableName} = '{text_content_add}'\n\n")


def image_cropper(screenshot_path, template_img_x1, template_img_y1, template_img_x2, template_img_y2):
    screenshot = p.screenshot()
    screenshot.save(screenshot_path)
   
    #mainImg,
    # delimiterx, delimitery = 25, 25
    #x1, y1, x2, y2 = xAxis - int(delimiterx), yAxis - int(delimitery), xAxis + int(delimiterx), yAxis + int(delimitery)
    # #print(x1, y1, x2, y2)
    x1, y1, x2, y2 = template_img_x1, template_img_y1, template_img_x2, template_img_y2
    if(x1 <= 0):
       x1 ,x2 = 0, x2 + -(x1)
    if(y1 <= 0):
       y1 ,y2 = 0, y2 + -(y1)
    if(x2 <= 0):
       x2, x1 = 0, x1 + -(x2)
    if(y2 <= 0):
       y2, y1  = 0, y1 + -(y2)
    if(x1 > 1919):
       x2, x1 = -(x1 - 1919 - x2), 1919
    if(y1 > 1079 ):
       y2, y1 = -(y1 - 1079 - y2), 1079
    if(x2 > 1919):
       x1, x2 = -(x2 - 1919 - x1), 1919
    if(y2 > 1079 ):
       y1, y2 = -(y2 - 1079 - y1), 1079
    print(x1, y1, x2, y2)
    # img = cv2.imread(mainImg)
    img = cv2.imread(screenshot_path)

    image_arr = np.array(img)      
    # Crop image
    image_arr = image_arr[y1:y2, x1:x2]    #250:350, 735:1174
    # Convert array to image
    image = Image.fromarray(image_arr)
    image.save(screenshot_path)
    # Resizing the image

def template_matching(mainimage, template_img, x1, y1, x2, y2):
    global xAxis
    global yAxis
    img = cv2.imread(mainimage,1)
    #cv2.imshow('Original',img)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    template = cv2.imread(template_img,0)
    #cv2.imshow('Template',template)
    w,h = template.shape[0], template.shape[1]

    matched = cv2.matchTemplate(gray,template,cv2.TM_CCOEFF_NORMED)
    threshold = 0.9

    loc = np.where( matched >= threshold)

    for pt in zip(*loc[::-1]):
        cv2.rectangle(img, pt, (pt[0] + w, pt[1] + h), (0,255,255), 2)
        #extra stuff if not needed close it.
        # cv2.imshow('Matched with Template',img)
        # print(x1 + pt[0], y1 + pt[1])
        # print(x1 + pt[0] + w, y1 + pt[1] + h)
    xAxis1 = x1 + pt[0]
    xAxis2 = x1 + pt[0] + w
    yAxis1 = y1 + pt[1]
    yAxis2 = y1 + pt[1] + h
    xAxis = (xAxis1 + xAxis2)/2
    yAxis = (yAxis1 + yAxis2)/2
    print(xAxis, yAxis)


def compare_words(str1, str2):
##    str1 = "aman54321"
##    str2 = "anbns432i"
    score = 0

    for i, (char1, char2) in enumerate(zip(str1, str2)):
        if char1 == char2:
            print(f"At index {i}, '{char1}' and '{char2}' are equal.")
            score = score + 1
        else:
            print(f"At index {i}, '{char1}' and '{char2}' are not equal.")

    # Handling different lengths after the loop
    if len(str1) != len(str2):
        print("Strings have different lengths.")

    print(score, len(str1))
    if(score > len(str1)/2):
        print("they are equal")


def loader_func(max_sec_time_out, x1,y1,x2,y2,key_text1,key_text2):
    max_v = max_sec_time_out * 2
    a = 0
    while a < max_v:
        ocrImageDetection(x1,y1,x2,y2,key_text1,key_text2)
        if(OCR_value == True):
            break
        t(0.5)
        a = a + 1

def loader_func_UI(max_sec_time_out, x1,y1,x2,y2,key_text1,key_text2):
    max_v = max_sec_time_out * 2
    a = 0
    while a < max_v:
        ocrImageDetectionUI(x1,y1,x2,y2,key_text1,key_text2)
        if(OCR_value == True):
            break
        t(0.5)
        a = a + 1

def temporary_excel(excel_file_name):
    import openpyxl
    global table_header_list
    global table_data_list
    file_path = r"C:\\Users\\Amanjyot\\Downloads"
    workbook = openpyxl.load_workbook(f'{file_path}\\{excel_file_name}.xlsx')
    sheet = workbook.active
    table_header_list = []
    table_data_list = []

##    #for creating a table header
##    for row in sheet.iter_rows():
##        for cell in row:
##            if cell.value == None:
##                break
##            row_number = cell.coordinate[1:]
##            table_header_list.append([cell.value, cell.col_idx, row_number])
##        break

    #for creating a table data
    for row in sheet.iter_rows():
        temp_table_data_list = []
        for cell in row:
            if cell.value == None:
                break
            row_number = cell.coordinate[1:]
            temp_table_data_list.append([cell.value, cell.col_idx, row_number])
        if temp_table_data_list != []:
            table_data_list.append(temp_table_data_list)
        
##    #for removing table header from table data
##    a = 0
##    while a < len(table_header_list):
##        if(table_header_list[a] in table_data_list):
##            table_data_list[0].remove(table_header_list[a])
##        a = a + 1

    table_header_list = table_data_list[0]
    del table_data_list[0]

    a = 0
    while a < len(table_data_list):
        insert_data(table_header_list, table_data_list, a)
        a = a + 1
        #break


def excel_reading(excel_file_name):
    import openpyxl
    global table_header_list
    global table_data_list
    a = 0
    while a < 200:
        check_if_it_got_downloaded(excel_file_name)
        if file_present == True:
            break
        t(0.5)
        a = a + 1
    if file_present == True:
        file_path = r"C:\\Users\\Amanjyot\\Downloads"
        workbook = openpyxl.load_workbook(f'{file_path}\\{excel_file_name}.xlsx')
        sheet = workbook.active
        table_header_list = []
        table_data_list = []

        #for creating a table data
        for row in sheet.iter_rows():
            temp_table_data_list = []
            for cell in row:
                if cell.value == None:
                    break
                row_number = cell.coordinate[1:]
                temp_table_data_list.append([cell.value, cell.col_idx, row_number])
            if temp_table_data_list != []:
                table_data_list.append(temp_table_data_list)

        table_header_list = table_data_list[0]
        del table_data_list[0]

        
        ocrImageDetectionUI(5,684,472,939,"START", "start")
        if(OCR_value == True):
            p.click(button='left',x=xAxis, y=yAxis)
            loader_func_UI(5,12,769,458,864,"ROUND", "round")
            if(OCR_value == True):
                a = 0
                while a < len(table_data_list):
                    insert_data(table_header_list, table_data_list, a)
                    a = a + 1
                    #break
        else:
            ocrImageDetectionUI(101,816,367,861,"START", "start")
            if(OCR_value == True):
                p.click(button='left',x=xAxis, y=yAxis)
                loader_func_UI(5,12,769,458,864,"ROUND", "round")
                if(OCR_value == True):
                    a = 0
                    while a < len(table_data_list):
                        insert_data(table_header_list, table_data_list, a)
                        a = a + 1
                        #break
            

def check_if_it_got_downloaded(excel_file_name):
    global file_present
    # print("check if the file got downloaded or not!")
    # Define the folder path and the file name
    file_path = r"C:\\Users\\Amanjyot\\Downloads"

    # Construct the full path to the file
    full_file_path = os.path.join(f'{file_path}\\{excel_file_name}.xlsx')

    # Use os.path.exists() to check for the file
    if os.path.exists(full_file_path):
        print(f"The file '{excel_file_name}' is present in the folder.")
        file_present = True
    else:
        print(f"The file '{excel_file_name}' is not present in the folder.")
        file_present = False

def insert_data(table_header_list, table_data_list, row_number):
    print(table_header_list, "\n", table_data_list, "\n", row_number)
    #to get all the information like email, last name etc input fields.
    global screen_coordinates_list
    screen_coordinates_list = []
    submit = ["SUBMIT", "submit"]
    continous_ocrImageDetection(522, 151, 1484 , 1037, "submit[0]", "submit[1]")
    #useless one
    #continous_ocrImageDetection(508, 91, 1376 , 1020, "submit[0]", "submit[1]")
    print(complete_text_array, "\n")
    for x in table_header_list:
        header_complete_text_array = []
        ax = re.split(r"\s", x[0])
        #print( "\n", ax)
        a = 0
        while a < len(complete_text_array):
            complete_text_array[a][0] = re.sub(r"\.|,|!|‘|-", "", complete_text_array[a][0])
            if complete_text_array[a][0] in ax:
                #print(complete_text_array[a][0], " : ", a)
                header_complete_text_array.append([complete_text_array[a][0], a])
            elif complete_text_array[a][0].casefold() in ax:
                #print(complete_text_array[a][0], " : ", a)
                header_complete_text_array.append([complete_text_array[a][0], a])
            a = a + 1
        print("header_complete_text_array", header_complete_text_array)
        a = 0
        while a < len(header_complete_text_array):
            if a + 1 < len(header_complete_text_array):
                current = header_complete_text_array[a]
                next_key = header_complete_text_array[a+ 1]
                print(current[0].lower() ,ax[0].lower() ,next_key[0].lower() , ax[1].lower())
                if current[0].lower() == ax[0].lower() and next_key[0].lower() == ax[1].lower() and current[1]+1 == next_key[1]:
                    #print(current, next_key)
                    actual_key = current[1]
                    screen_coordinates_list.append(complete_text_array[actual_key])
                    break
            else:
                #actual_key = header_complete_text_array[a]
                #print(actual_key)
                actual_key = header_complete_text_array[a][1]
                screen_coordinates_list.append(complete_text_array[actual_key])
            a = a + 1
    screen_coordinates_list.append(complete_text_array[-1])
    print(complete_text_array, "\n" ,screen_coordinates_list)

    #inserting_actual_data
    a = 0
    while a < len(screen_coordinates_list)-1:
        if a < len(table_data_list[row_number]):
            value = table_data_list[row_number][a][0]
            p.click(button='left',x=screen_coordinates_list[a][1]+20, y=screen_coordinates_list[a][2]+40)
            p.write(str(value))
            p.press("esc")
        a = a + 1
    #click on submit button
    if screen_coordinates_list[-1][0] == "SUBMIT" or screen_coordinates_list[-1][0] == "JBMIT" or screen_coordinates_list[-1][0] == "submit":
        p.moveTo(x=screen_coordinates_list[-1][1], y=screen_coordinates_list[-1][2])
    else:
        ocrImageDetection(538, 291, 823 , 1006, "SUBMIT" , "JBMIT")
        if(OCR_value == True):
            p.moveTo(x=xAxis, y=yAxis)
        else:
            ocrImageDetectionUI(538, 291, 823 , 1006, "SUBMIT" , "JBMIT")
            if(OCR_value == True):
                p.moveTo(x=xAxis, y=yAxis)
    
    p.click()


def check_teams_messages():
    global process
    process = "False"
    ocrImageDetection(577,313,876,367,'ajay','shankar')
    if(OCR_value == True):
        loader_func(300, 566,895,1493,952,'start','start')
        if(OCR_value == True):
            process = "True"
        else:
            loader_func(300, 566,895,1493,952,'stop','stop')
            if(OCR_value == True):
                process = "False"
            
    







### Main function starts from here onwards ###
global excel_file_name
def input_data_challenge(excel_file_name):
    URL = "https://rpachallenge.azurewebsites.net/"
    #Open chrome
    p.click(button='left',x=458, y=1063)
    #search bar for google
    loader_func(3, 63 ,8,126 ,37,"New","tab")
    if(OCR_value == True):
        # p.click(button='left',x=xAxis, y=yAxis)        # t(0.5)        # ocrImageDetection(149,46,341,81,'type','URL')        # if(OCR_value == True):        # p.click(button='left',x=xAxis, y=yAxis)        # t(0.5)        #type URL = "https://rpachallenge.azurewebsites.net/"
        p.write(URL)        #hit enter key
        p.press('esc')
        p.press('enter')        # wait for site to be loaded.
    loader_func(10,5,87,299,144,'RPA','Challenge')
    if(OCR_value == True):
        loader_func_UI(5,2,630,471,937,'DOWNLOAD','EXCEL')          #wait for download excel to appear
        if(OCR_value == True):
            t(0.5)
            p.click(button='left',x=xAxis, y=yAxis)
            loader_func_UI(5,63 ,410,920 ,449,'File','name')          #wait for excel file to appear
            if(OCR_value == True):
                p.write(excel_file_name)
                p.press("enter")
                # wait for duplicate error if appears.
                loader_func_UI(5,817 ,453,1002 ,506,'already','replace')
                if(OCR_value == True):
                    ocrImageDetectionUI(977,520,1039,538,'Yes','yes')
                    if(OCR_value == True):
                        p.click(button='left',x=xAxis, y=yAxis)
                    else:
                        p.press("tab")
                        p.press("enter")
                    t(1)
                # question how to wait till the file is downloaded? how to do it?
                excel_reading(excel_file_name)
##                insert_data(table_header_list, table_data_list, row_number)
##                print("::", screen_coordinates_list)



       
        







def OCR_challenge():
    URL2 = "https://rpachallengeocr.azurewebsites.net/"
    invoice_download_icon = (r'C:\Users\Amanjyot\Desktop\python\ultimate_bot\small_images\invoice_download_icon.png', r'C:\Users\Amanjyot\Desktop\small_screenshot.png', 1676,338,1693,354)
    x1, y1, x2, y2 = 1676,338,1693,354
    template_img_path = r'C:\Users\Amanjyot\Desktop\python\automation_realworld_examples\icons\invoice_download_icon.png'
    main_image_path = r'C:\Users\Amanjyot\Desktop\small_screenshot.png'
    image_cropper(main_image_path, x1, y1, x2, y2)
    template_matching(main_image_path, template_img_path, x1, y1, x2, y2)
    """Download
    example
    csv
    &,
    Download
    sample
    invoice
    1
    &,
    Download
    sample
    invoice
    2"""













##table_header_list = [['First Name', 'A1'], ['Last Name ', 'B1'], ['Company Name', 'C1'], ['Role in Company', 'D1'], ['Address', 'E1'], ['Email', 'F1'], ['Phone Number', 'G1']]
##table_data_list = [['John', 'A2'], ['Smith', 'B2'], ['IT Solutions', 'C2'], ['Analyst', 'D2'], ['98 North Road', 'E2'], ['jsmith@itsolutions.co.uk', 'F2'], [40716543298, 'G2'], ['Jane', 'A3'], ['Dorsey', 'B3'], ['MediCare', 'C3'], ['Medical Engineer', 'D3'], ['11 Crown Street', 'E3'], ['jdorsey@mc.com', 'F3'], [40791345621, 'G3'], ['Albert', 'A4'], ['Kipling', 'B4'], ['Waterfront', 'C4'], ['Accountant', 'D4'], ['22 Guild Street', 'E4'], ['kipling@waterfront.com', 'F4'], [40735416854, 'G4'], ['Michael', 'A5'], ['Robertson', 'B5'], ['MediCare', 'C5'], ['IT Specialist', 'D5'], ['17 Farburn Terrace', 'E5'], ['mrobertson@mc.com', 'F5'], [40733652145, 'G5'], ['Doug', 'A6'], ['Derrick', 'B6'], ['Timepath Inc.', 'C6'], ['Analyst', 'D6'], ['99 Shire Oak Road', 'E6'], ['dderrick@timepath.co.uk', 'F6'], [40799885412, 'G6'], ['Jessie', 'A7'], ['Marlowe', 'B7'], ['Aperture Inc.', 'C7'], ['Scientist', 'D7'], ['27 Cheshire Street', 'E7'], ['jmarlowe@aperture.us', 'F7'], [40733154268, 'G7'], ['Stan', 'A8'], ['Hamm', 'B8'], ['Sugarwell', 'C8'], ['Advisor', 'D8'], ['10 Dam Road', 'E8'], ['shamm@sugarwell.org', 'F8'], [40712462257, 'G8'], ['Michelle', 'A9'], ['Norton', 'B9'], ['Aperture Inc.', 'C9'], ['Scientist', 'D9'], ['13 White Rabbit Street', 'E9'], ['mnorton@aperture.us', 'F9'], [40731254562, 'G9'], ['Stacy', 'A10'], ['Shelby', 'B10'], ['TechDev', 'C10'], ['HR Manager', 'D10'], ['19 Pineapple Boulevard', 'E10'], ['sshelby@techdev.com', 'F10'], [40741785214, 'G10'], ['Lara', 'A11'], ['Palmer', 'B11'], ['Timepath Inc.', 'C11'], ['Programmer', 'D11'], ['87 Orange Street', 'E11'], ['lpalmer@timepath.co.uk', 'F11'], [40731653845, 'G11']]
##p.click(x=1657, y=247)
##temporary_excel("challenge")
# insert_data(table_header_list, table_data_list, row_number)
# excel_reading("challenge")


#facing problem here fix it later.
##process = False
##while process == False:
##    check_teams_messages()
##    print(process)
##    if process == True:
##        input_data_challenge("challenge")
##    t(0.5)



input_data_challenge("challenge")
